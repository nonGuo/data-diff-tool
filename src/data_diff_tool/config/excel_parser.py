"""Excel parser - reads entity and attribute mapping sheets to generate verification tasks."""

from __future__ import annotations

import logging
from collections import defaultdict
from typing import Any

import openpyxl

from data_diff_tool.config.models import (
    Column,
    ColumnMapping,
    EntityMapping,
    InventoryTask,
    SkippedTask,
    VerificationTask,
)

logger = logging.getLogger(__name__)

ENTITY_SHEET = "实体级mapping"
ATTRIBUTE_SHEET = "属性级mapping"
SAMPLE_SHEET = "抽样校验配置"

# ── 实体级 mapping 列索引 (0-based) ──────────────────────────────
# 序号(0) | 切换前库名(1) | 切换前Schema(2) | 切换前表名(3)
# | 切换后库名(4) | 切换后Schema(5) | 切换后表名(6) | 实体级变化类型(7)
# | 数据迁移策略(8) | 迁移后粒度是否发生变化(9) | 详细说明(10)
ENT_OLD_DB = 1
ENT_OLD_SCHEMA = 2
ENT_OLD_TABLE = 3
ENT_NEW_DB = 4
ENT_NEW_SCHEMA = 5
ENT_NEW_TABLE = 6
ENT_MAPPING_TYPE = 7
ENT_DESCRIPTION = 10

# ── 抽样校验配置 列索引 (0-based) ────────────────────────────────
# 序号(0) | 切换前库名(1) | 切换前Schema(2) | 切换前表名(3)
# | 主键字段(4) | 过滤条件(5) | 备注(6)
SMP_OLD_DB = 1
SMP_OLD_SCHEMA = 2
SMP_OLD_TABLE = 3
SMP_PRIMARY_KEYS = 4
SMP_FILTER_COND = 5
SMP_REMARK = 6

# ── 属性级 mapping 列索引 (0-based) ──────────────────────────────
# 序号(0) | 切换前库名(1) | 切换前Schema(2) | 切换前表名(3)
# | 切换前字段名(4) | 切换前字段中文名(5) | 切换前字段类型(6)
# | 切换后库名(7) | 切换后Schema(8) | 切换后表名(9)
# | 切换后字段名(10) | 切换后字段中文名(11) | 切换后字段类型(12)
# | 字段级变化类型(13) | 数据内容变化(14) | 是否可还原(15) | 还原方案详细说明(16)
ATT_OLD_DB = 1
ATT_OLD_SCHEMA = 2
ATT_OLD_TABLE = 3
ATT_OLD_COLUMN = 4
ATT_OLD_TYPE = 6
ATT_NEW_DB = 7
ATT_NEW_SCHEMA = 8
ATT_NEW_TABLE = 9
ATT_NEW_COLUMN = 10
ATT_NEW_TYPE = 12
ATT_CHANGE_TYPE = 13
ATT_DATA_CHANGED = 14


def _make_fqn(db: str | None, schema: str | None, table: str | None) -> str:
    """Build a fully qualified name from parts, returning empty string if table is missing."""
    db = (db or "").strip()
    schema = (schema or "").strip()
    table = (table or "").strip()
    if not table:
        return ""
    parts = [p for p in (db, schema, table) if p]
    return ".".join(parts)


def _is_data_changed(raw: Any) -> bool:
    """Parse the '数据内容变化' column to determine if data has changed."""
    if not raw:
        return False
    val = str(raw).strip()
    # 文档中示例: "1.数据内容不变" / "2.数据值域变化"
    # 包含"不变"表示数据内容未变化，其余视为有变化
    if "不变" in val:
        return False
    return bool(val)


class ExcelParser:
    """Parses entity and attribute mapping sheets from Excel files."""

    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def parse(
        self,
        *,
        primary_keys: list[str] | None = None,
        filter_cond: str | None = None,
    ) -> list[VerificationTask | SkippedTask | InventoryTask]:
        """
        Parse the Excel file and return a list of tasks.

        Args:
            primary_keys: Optional list of primary key column names as global fallback.
                Per-table settings in the sample sheet take priority.
            filter_cond: Optional WHERE filter condition as global fallback.
        """
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)

        if ENTITY_SHEET not in wb.sheetnames:
            raise ValueError(f"Sheet '{ENTITY_SHEET}' not found in {self.file_path}")
        if ATTRIBUTE_SHEET not in wb.sheetnames:
            raise ValueError(f"Sheet '{ATTRIBUTE_SHEET}' not found in {self.file_path}")

        # Parse entity mapping
        entity_mappings = self._parse_entity_sheet(wb[ENTITY_SHEET])

        # Parse attribute mapping and group by entity key
        attr_by_entity = self._parse_attribute_sheet(wb[ATTRIBUTE_SHEET])

        # Parse sample configuration (optional sheet)
        sample_config: dict[str, tuple[list[str], str | None]] = {}
        if SAMPLE_SHEET in wb.sheetnames:
            sample_config = self._parse_sample_sheet(wb[SAMPLE_SHEET])

        tasks = self._build_tasks(
            entity_mappings,
            attr_by_entity,
            sample_config=sample_config,
            fallback_primary_keys=primary_keys or [],
            fallback_filter_cond=filter_cond,
        )

        wb.close()
        return tasks

    # ── Entity sheet parsing ──────────────────────────────────────

    def _parse_entity_sheet(self, sheet: Any) -> list[EntityMapping]:
        """Parse the entity-level mapping sheet into EntityMapping objects."""
        entities: list[EntityMapping] = []
        row_num = 1  # header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_num += 1
            mapping_type = (row[ENT_MAPPING_TYPE] or "").strip()
            if not mapping_type:
                logger.debug("Row %d: skipping (no mapping type)", row_num)
                continue

            old_fqn = _make_fqn(row[ENT_OLD_DB], row[ENT_OLD_SCHEMA], row[ENT_OLD_TABLE])
            new_fqn = _make_fqn(row[ENT_NEW_DB], row[ENT_NEW_SCHEMA], row[ENT_NEW_TABLE])
            description = (row[ENT_DESCRIPTION] or "").strip()

            entities.append(EntityMapping(
                old_fqn=old_fqn,
                new_fqn=new_fqn,
                mapping_type=mapping_type,
                description=description,
            ))
            logger.debug(
                "Row %d: %s -> %s (%s)",
                row_num, old_fqn or "(none)", new_fqn or "(none)", mapping_type,
            )
        return entities

    # ── Attribute sheet parsing ───────────────────────────────────

    def _parse_attribute_sheet(
        self,
        sheet: Any,
    ) -> dict[tuple[str, str], list[ColumnMapping]]:
        """
        Parse the attribute-level mapping sheet.

        Returns a dict mapping (old_fqn, new_fqn) -> list of ColumnMapping.
        """
        grouped: dict[tuple[str, str], list[ColumnMapping]] = defaultdict(list)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            change_type = (row[ATT_CHANGE_TYPE] or "").strip()
            if not change_type:
                continue

            old_fqn = _make_fqn(row[ATT_OLD_DB], row[ATT_OLD_SCHEMA], row[ATT_OLD_TABLE])
            new_fqn = _make_fqn(row[ATT_NEW_DB], row[ATT_NEW_SCHEMA], row[ATT_NEW_TABLE])

            old_col_name = (row[ATT_OLD_COLUMN] or "").strip()
            new_col_name = (row[ATT_NEW_COLUMN] or "").strip()
            if not old_col_name and not new_col_name:
                continue

            old_col = Column(name=old_col_name, data_type=(row[ATT_OLD_TYPE] or "").strip())
            new_col = Column(name=new_col_name, data_type=(row[ATT_NEW_TYPE] or "").strip())

            data_changed = _is_data_changed(row[ATT_DATA_CHANGED])

            cm = ColumnMapping(
                old_fqn=old_fqn,
                new_fqn=new_fqn,
                old_col=old_col,
                new_col=new_col,
                change_type=change_type,
                data_changed=data_changed,
            )

            entity_key = (old_fqn, new_fqn)
            grouped[entity_key].append(cm)

        return grouped

    # ── Sample configuration parsing ──────────────────────────────

    def _parse_sample_sheet(
        self,
        sheet: Any,
    ) -> dict[str, tuple[list[str], str | None]]:
        """
        Parse the sample verification configuration sheet.

        Returns a dict mapping old_fqn -> (primary_keys, filter_cond).
        """
        config: dict[str, tuple[list[str], str | None]] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            old_fqn = _make_fqn(row[SMP_OLD_DB], row[SMP_OLD_SCHEMA], row[SMP_OLD_TABLE])
            if not old_fqn:
                continue

            # Parse primary keys (comma-separated)
            pk_raw = (row[SMP_PRIMARY_KEYS] or "").strip()
            primary_keys = [k.strip() for k in pk_raw.split(",") if k.strip()] if pk_raw else []

            # Parse filter condition
            filter_raw = (row[SMP_FILTER_COND] or "").strip()
            filter_cond = filter_raw if filter_raw else None

            config[old_fqn] = (primary_keys, filter_cond)
            logger.debug(
                "Sample config for %s: pks=%s, filter=%s",
                old_fqn, primary_keys, filter_cond,
            )

        return config

    # ── Task building ─────────────────────────────────────────────

    def _build_tasks(
        self,
        entities: list[EntityMapping],
        attr_by_entity: dict[tuple[str, str], list[ColumnMapping]],
        sample_config: dict[str, tuple[list[str], str | None]],
        *,
        fallback_primary_keys: list[str],
        fallback_filter_cond: str | None,
    ) -> list[VerificationTask | SkippedTask | InventoryTask]:
        """Build task objects from parsed mappings, distributing column info to 1:1 tasks."""
        tasks: list[VerificationTask | SkippedTask | InventoryTask] = []

        for entity in entities:
            key = entity.composite_key
            col_mappings = attr_by_entity.get(key, [])

            if entity.mapping_type == "1:1":
                # Get per-table config from sample sheet, fallback to CLI/global defaults
                pks, fcond = sample_config.get(entity.old_fqn, ([], None))
                if not pks:
                    pks = fallback_primary_keys
                if not fcond:
                    fcond = fallback_filter_cond
                task = self._build_verification_task(entity, col_mappings, pks, fcond)
                tasks.append(task)
            elif entity.mapping_type in ("1:N", "N:1"):
                tasks.append(SkippedTask(
                    entity=entity,
                    reason=f"Mapping type {entity.mapping_type} requires manual review",
                ))
            elif entity.mapping_type == "1:0":
                tasks.append(InventoryTask(entity=entity))
            elif entity.mapping_type == "0:1":
                tasks.append(InventoryTask(entity=entity))
            else:
                raise ValueError(f"Unknown mapping type '{entity.mapping_type}' for {key}")

        return tasks

    def _build_verification_task(
        self,
        entity: EntityMapping,
        col_mappings: list[ColumnMapping],
        primary_keys: list[str],
        filter_cond: str | None,
    ) -> VerificationTask:
        """
        Build a VerificationTask for a 1:1 entity mapping.

        Column classification rules:
        - "1:1 完全一致"              → identical_columns
        - "1:1 字段类型变化"           → cast_columns
        - "1:1 数据内容变化"           → skipped_columns (data changed, skip comparison)
        - "1:1 字段类型变化 数据内容变化" → skipped_columns
        """
        identical: list[str] = []
        cast_cols: list[str] = []
        skipped: list[str] = []

        for cm in col_mappings:
            # Determine which column name to use (prefer old column name)
            col_name = cm.old_col.name or cm.new_col.name
            ct = cm.change_type

            if "完全一致" in ct:
                identical.append(col_name)
            elif "字段类型变化" in ct and not cm.data_changed:
                cast_cols.append(col_name)
            else:
                # "数据内容变化" or "字段类型变化 数据内容变化" or unknown
                skipped.append(col_name)
                logger.debug(
                    "Skipping column '%s' for %s (change_type='%s', data_changed=%s)",
                    col_name, entity.composite_key, ct, cm.data_changed,
                )

        return VerificationTask(
            entity=entity,
            primary_keys=primary_keys,
            identical_columns=identical,
            cast_columns=cast_cols,
            skipped_columns=skipped,
            filter_cond=filter_cond,
        )
