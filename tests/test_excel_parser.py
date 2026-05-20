"""Tests for Excel parser."""

import os
import pytest
import openpyxl

from data_diff_tool.config.excel_parser import ExcelParser
from data_diff_tool.config.models import (
    VerificationTask,
    SkippedTask,
    InventoryTask,
)

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
os.makedirs(FIXTURE_DIR, exist_ok=True)

SAMPLE_XLSX = os.path.join(FIXTURE_DIR, "test_mapping.xlsx")
SAMPLE_WITH_CONFIG_XLSX = os.path.join(FIXTURE_DIR, "test_mapping_with_sample.xlsx")
SAMPLE_WITHOUT_SHEET_XLSX = os.path.join(FIXTURE_DIR, "test_mapping_no_sample.xlsx")


def _create_test_excel(path: str, include_sample_sheet: bool = False) -> None:
    """Helper to create a test Excel file."""
    wb = openpyxl.Workbook()

    ws_entity = wb.active
    ws_entity.title = "实体级mapping"
    ws_entity.append([
        "序号", "切换前库名", "切换前Schema", "切换前表名",
        "切换后库名", "切换后Schema", "切换后表名",
        "实体级变化类型", "数据迁移策略", "迁移后粒度是否发生变化", "详细说明"
    ])
    ws_entity.append([1, "edw", "sdi", "sdi_contract_2000", "edw", "sdi", "sdi_contract_3000", "1:1", "全量迁移", "否", ""])
    ws_entity.append([2, "edw", "sdi", "sdi_order_2000", "edw", "sdi", "sdi_order_cn_3000", "1:N", "全量迁移", "否", ""])
    ws_entity.append([3, "edw", "sdi", "sdi_user_2000", "", "", "", "1:0", "", "", ""])
    ws_entity.append([4, "", "", "", "edw", "sdi", "sdi_partner_3000", "0:1", "", "", ""])

    ws_attr = wb.create_sheet("属性级mapping")
    ws_attr.append([
        "序号", "切换前库名", "切换前Schema", "切换前表名", "切换前字段名", "切换前字段中文名", "切换前字段类型",
        "切换后库名", "切换后Schema", "切换后表名", "切换后字段名", "切换后字段中文名", "切换后字段类型",
        "字段级变化类型", "数据内容变化", "是否可还原", "还原方案详细说明"
    ])
    ws_attr.append([
        1, "edw", "sdi", "sdi_contract_2000", "contract_id", "合同ID", "nvarchar2(100)",
        "edw", "sdi", "sdi_contract_3000", "contract_id", "合同ID", "nvarchar2(100)",
        "1:1 完全一致", "1.数据内容不变", "", ""
    ])
    ws_attr.append([
        2, "edw", "sdi", "sdi_contract_2000", "contract_no", "合同号", "nvarchar2(100)",
        "edw", "sdi", "sdi_contract_3000", "contract_num", "合同号", "nvarchar2(200)",
        "1:1 字段类型变化", "1.数据内容不变", "", ""
    ])
    ws_attr.append([
        3, "edw", "sdi", "sdi_contract_2000", "contract_type", "合同类型", "nvarchar2(30)",
        "edw", "sdi", "sdi_contract_3000", "contract_type", "合同类型", "nvarchar2(30)",
        "1:1 数据内容变化", "2.数据值域变化", "Y", ""
    ])

    if include_sample_sheet:
        ws_sample = wb.create_sheet("抽样校验配置")
        ws_sample.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "主键字段", "过滤条件", "备注"
        ])
        ws_sample.append([
            1, "edw", "sdi", "sdi_contract_2000", "contract_id", "dt = '2026-03-01'", ""
        ])

    wb.save(path)


@pytest.fixture(scope="module", autouse=True)
def create_test_files():
    """Create test Excel files for all tests."""
    _create_test_excel(SAMPLE_XLSX, include_sample_sheet=False)
    _create_test_excel(SAMPLE_WITH_CONFIG_XLSX, include_sample_sheet=True)
    yield


class TestParseEntitySheet:
    def test_parse_1_1_entity(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task_1_1 = tasks[0]
        assert isinstance(task_1_1, VerificationTask)
        assert task_1_1.entity.old_fqn == "edw.sdi.sdi_contract_2000"
        assert task_1_1.entity.new_fqn == "edw.sdi.sdi_contract_3000"
        assert task_1_1.entity.mapping_type == "1:1"

    def test_parse_1_N_entity(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task_1n = tasks[1]
        assert isinstance(task_1n, SkippedTask)
        assert task_1n.entity.mapping_type == "1:N"

    def test_parse_1_0_entity(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task_1_0 = tasks[2]
        assert isinstance(task_1_0, InventoryTask)
        assert task_1_0.entity.old_fqn == "edw.sdi.sdi_user_2000"
        assert task_1_0.entity.new_fqn == ""

    def test_parse_0_1_entity(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task_0_1 = tasks[3]
        assert isinstance(task_0_1, InventoryTask)
        assert task_0_1.entity.old_fqn == ""
        assert task_0_1.entity.new_fqn == "edw.sdi.sdi_partner_3000"


class TestColumnClassification:
    def test_identical_columns(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert "contract_id" in task.identical_columns

    def test_cast_columns(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert "contract_no" in task.cast_columns

    def test_skipped_columns(self):
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert "contract_type" in task.skipped_columns


class TestSampleSheetParsing:
    def test_sample_config_applied(self):
        """Test that 抽样校验配置 sheet values are applied to 1:1 tasks."""
        parser = ExcelParser(SAMPLE_WITH_CONFIG_XLSX)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert task.primary_keys == ["contract_id"]
        assert task.filter_cond == "dt = '2026-03-01'"

    def test_sample_config_composite_key(self):
        """Test composite primary key parsing."""
        # Create temp excel with composite key
        wb = openpyxl.Workbook()
        ws_entity = wb.active
        ws_entity.title = "实体级mapping"
        ws_entity.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "切换后库名", "切换后Schema", "切换后表名",
            "实体级变化类型", "数据迁移策略", "迁移后粒度是否发生变化", "详细说明"
        ])
        ws_entity.append([1, "edw", "sdi", "t1", "edw", "sdi", "t2", "1:1", "", "", ""])

        ws_attr = wb.create_sheet("属性级mapping")
        ws_attr.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名", "切换前字段名", "切换前字段中文名", "切换前字段类型",
            "切换后库名", "切换后Schema", "切换后表名", "切换后字段名", "切换后字段中文名", "切换后字段类型",
            "字段级变化类型", "数据内容变化", "是否可还原", "还原方案详细说明"
        ])
        ws_attr.append([
            1, "edw", "sdi", "t1", "id", "ID", "int8",
            "edw", "sdi", "t2", "id", "ID", "int8",
            "1:1 完全一致", "1.数据内容不变", "", ""
        ])

        ws_sample = wb.create_sheet("抽样校验配置")
        ws_sample.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "主键字段", "过滤条件", "备注"
        ])
        ws_sample.append([
            1, "edw", "sdi", "t1", "col_a,col_b", "status = 'active'", ""
        ])

        path = os.path.join(FIXTURE_DIR, "test_composite_pk.xlsx")
        wb.save(path)

        parser = ExcelParser(path)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert task.primary_keys == ["col_a", "col_b"]
        assert task.filter_cond == "status = 'active'"

    def test_cli_fallback_when_no_sample_sheet(self):
        """Test that CLI --primary-keys and --filter are used as fallback."""
        parser = ExcelParser(SAMPLE_XLSX)
        tasks = parser.parse(
            primary_keys=["contract_id"],
            filter_cond="dt = '2026-01-01'",
        )
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert task.primary_keys == ["contract_id"]
        assert task.filter_cond == "dt = '2026-01-01'"

    def test_sample_sheet_overrides_cli_fallback(self):
        """Test that sample sheet values override CLI fallbacks."""
        parser = ExcelParser(SAMPLE_WITH_CONFIG_XLSX)
        tasks = parser.parse(
            primary_keys=["global_pk"],
            filter_cond="global_filter",
        )
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        # Sample sheet values should win over CLI fallback
        assert task.primary_keys == ["contract_id"]
        assert task.filter_cond == "dt = '2026-03-01'"

    def test_empty_sample_sheet_row(self):
        """Test that empty rows in sample sheet are skipped."""
        wb = openpyxl.Workbook()
        ws_entity = wb.active
        ws_entity.title = "实体级mapping"
        ws_entity.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "切换后库名", "切换后Schema", "切换后表名",
            "实体级变化类型", "数据迁移策略", "迁移后粒度是否发生变化", "详细说明"
        ])
        ws_entity.append([1, "edw", "sdi", "t1", "edw", "sdi", "t2", "1:1", "", "", ""])

        ws_attr = wb.create_sheet("属性级mapping")
        ws_attr.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名", "切换前字段名", "切换前字段中文名", "切换前字段类型",
            "切换后库名", "切换后Schema", "切换后表名", "切换后字段名", "切换后字段中文名", "切换后字段类型",
            "字段级变化类型", "数据内容变化", "是否可还原", "还原方案详细说明"
        ])

        ws_sample = wb.create_sheet("抽样校验配置")
        ws_sample.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "主键字段", "过滤条件", "备注"
        ])
        # Empty row
        ws_sample.append([1, "", "", "", "", "", ""])
        # Valid row
        ws_sample.append([2, "edw", "sdi", "t1", "id", "", ""])

        path = os.path.join(FIXTURE_DIR, "test_empty_sample_row.xlsx")
        wb.save(path)

        parser = ExcelParser(path)
        tasks = parser.parse()
        task = tasks[0]
        assert isinstance(task, VerificationTask)
        assert task.primary_keys == ["id"]
        assert task.filter_cond is None


class TestEdgeCases:
    def test_missing_entity_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.create_sheet("属性级mapping")
        path = str(tmp_path / "bad.xlsx")
        wb.save(path)
        parser = ExcelParser(path)
        with pytest.raises(ValueError, match="实体级mapping"):
            parser.parse()

    def test_empty_attribute_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "实体级mapping"
        wb.active.append([
            "序号", "切换前库名", "切换前Schema", "切换前表名",
            "切换后库名", "切换后Schema", "切换后表名",
            "实体级变化类型", "数据迁移策略", "迁移后粒度是否发生变化", "详细说明"
        ])
        wb.active.append([1, "edw", "sdi", "t1", "edw", "sdi", "t2", "1:1", "", "", ""])
        wb.create_sheet("属性级mapping")
        path = str(tmp_path / "empty_attr.xlsx")
        wb.save(path)
        parser = ExcelParser(path)
        tasks = parser.parse()
        assert len(tasks) == 1
        assert isinstance(tasks[0], VerificationTask)
        assert tasks[0].identical_columns == []
